# -*- coding: utf-8 -*-
# Karlik xlsx -> scripts/karlik-import.json  (Produkte + Farbvarianten)
import openpyxl, json, re
from collections import defaultdict, Counter

SRC = "C:/Users/alisa/Downloads/IMERA_Karlik_Ali_Import.xlsx"
OUT = "scripts/karlik-import.json"

def slugify(s):
    s = str(s or '').lower()
    s = s.replace('ä','ae').replace('ö','oe').replace('ü','ue').replace('ß','ss')
    s = re.sub(r'[^a-z0-9]+','-', s).strip('-')
    return s

CAT_MAP = {
    '⚠️ prüfen: frames - glass effect - DECO Art': 'Rahmen Glasoptik DECO Art',
    '⚠️ prüfen: modular junction boxes': 'Modulare Verbindungsdosen',
    '⚠️ prüfen: DECO Pastel Matt': 'Rahmen DECO Pastell Matt',
    '⚠️ prüfen: dimmers': 'Dimmer',
    '⚠️ prüfen: frames - glass effect - non-standard modular': 'Rahmen Glasoptik modular (Sondermaße)',
    '⚠️ prüfen: central vacum cleaner suction socket': 'Zentralstaubsauger-Dose',
}

# Farb-Reihenfolge (Weiß zuerst, dann gängige Farben)
COLOR_ORDER = ['Weiß','Mattweiß','Beige','Silber Metallic','Gold','Gold Metallic','Graphit',
               'Mattgraphit','Mattgrau','Mattschwarz','Braun Metallic','Taupe','Salbeigrün',
               'Lachs','Terrakotta']
def color_rank(c):
    return COLOR_ORDER.index(c) if c in COLOR_ORDER else len(COLOR_ORDER) + (hash(c) % 100)

def clean(s):
    return str(s or '').replace('\xa0',' ').strip()

def strip_color_paren(n):
    return re.sub(r'\s*\([^)]*\)\s*$','', n).strip()

def base_name(vs):
    # Bevorzugt deutscher Name ohne ⚠️; sonst Englisch (Spalte D)
    ger = [strip_color_paren(clean(v[2])) for v in vs if not (v[2] and '⚠' in str(v[2]))]
    if ger:
        return Counter(ger).most_common(1)[0][0], False
    en = clean(vs[0][3]) or clean(vs[0][2])
    en = re.sub(r'^⚠️?\s*manuell prüfen:\s*','', en, flags=re.I)
    return en, True

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb["Import"]
data = [r for i,r in enumerate(ws.iter_rows(values_only=True)) if i>=5 and r[0] is not None]

# Gruppieren nach (Basis-Artikel, Serie)
groups = defaultdict(list)
for r in data:
    groups[(r[1], r[4])].append(r)

cats = {}   # slug -> name
products = []
sku_used = set()
slug_used = set()
warn_products = []

for (base, serie), vs in groups.items():
    # Kategorie (häufigste in der Gruppe), ⚠️ gemappt
    cat_raw = Counter(clean(v[5]) for v in vs).most_common(1)[0][0]
    cat_name = CAT_MAP.get(cat_raw, cat_raw)
    cat_slug = slugify(cat_name)
    cats.setdefault(cat_slug, cat_name)

    name, warned = base_name(vs)
    # eindeutige SKU/slug
    sku = str(base)
    if sku in sku_used:
        sku = f"{base}-{serie}"
    sku_used.add(sku)
    slug = slugify(f"{base}-{serie}")
    if slug in slug_used:
        n = 2
        while f"{slug}-{n}" in slug_used: n += 1
        slug = f"{slug}-{n}"
    slug_used.add(slug)

    # Varianten sortiert
    variants = []
    for v in sorted(vs, key=lambda x: (color_rank(clean(x[6])), clean(x[6]))):
        variants.append({
            "sku": clean(v[0]),
            "color": clean(v[6]),
            "ean": clean(v[7]),
            "price": round(float(v[8]), 2) if v[8] is not None else None,
            "image": clean(v[10]),
        })
    variants = [x for x in variants if x["price"] is not None]
    if not variants:
        continue
    price_min = min(x["price"] for x in variants)
    default = variants[0]  # Weiß bzw. erste

    if warned:
        warn_products.append(sku)

    products.append({
        "sku": sku,
        "slug": slug,
        "name": name,
        "series": serie,
        "category_slug": cat_slug,
        "short_description": f"Karlik {serie} – {cat_name}",
        "image": default["image"],
        "price_min": price_min,
        "warn": warned,
        "variants": variants,
    })

# ─── American-Style-Schalter aus der Original-Preisliste (Kipphebel-Ausführung) ───
# Nachtrag: technisch identisch mit Standard-Schaltern, nur Kipphebel statt Wippe.
# Preis = Karlik-Listenpreis (CENA NETTO EUR) × 0,78. Name = Standard-Name + " (Kipphebel-Ausführung)".
import os as _os
def _find_orig():
    d = "C:/Users/alisa/Downloads"
    for f in _os.listdir(d):
        if 'cennik' in f.lower():
            return _os.path.join(d, f)
    return None

# Standard-Name/-Kategorie je (Symbol, Serie) aus der Import-Datei
name_by, cat_by = {}, {}
for r in data:
    name_by[(str(r[0]), r[4])] = clean(r[2])
    cat_by[(str(r[0]), r[4])] = clean(r[5])

EN_DE_COLOR = {
    'white':'Weiß','matt white':'Mattweiß','beige':'Beige','silver metallic':'Silber Metallic',
    'gold':'Gold','gold metallic':'Gold Metallic','graphite':'Graphit','matt graphite':'Mattgraphit',
    'matt grey':'Mattgrau','matt black':'Mattschwarz','brown metallic':'Braun Metallic','taupe':'Taupe',
    'sage':'Salbeigrün','salmon':'Lachs','terracotta':'Terrakotta','navy blue':'Marineblau',
}
def core_sym(sym):
    m = re.match(r'^(\d*)([A-Za-z].*)$', str(sym))
    return m.group(2) if m else str(sym)

orig = _find_orig()
ame_added = 0
if orig:
    ows = openpyxl.load_workbook(orig, read_only=True, data_only=True)['CENNIK_PRICE LIST_ПРАЙС-ЛИСТ']
    orows = [r for r in ows.iter_rows(min_row=8, values_only=True) if r[1]]
    ame = [r for r in orows if r[18] and 'american' in str(r[18]).lower()]
    agroups = defaultdict(list)
    for r in ame:
        agroups[(core_sym(r[1]), r[6])].append(r)
    for (cs, serie), vs in agroups.items():
        ss = cs.replace('US', '', 1)
        std = name_by.get((ss, serie))
        if std:
            nm = re.sub(r'^⚠️?\s*manuell prüfen:\s*', '', std, flags=re.I).strip()
            name = f"{nm} (Kipphebel-Ausführung)"
        else:
            name = f"{clean(vs[0][13] or vs[0][12])} (Kipphebel-Ausführung)"
        cat_name = cat_by.get((ss, serie), 'Schalter') or 'Schalter'
        cat_slug = slugify(cat_name); cats.setdefault(cat_slug, cat_name)
        sku = cs if cs not in sku_used else f"{cs}-{serie}"
        sku_used.add(sku)
        slug = slugify(f"{cs}-{serie}")
        if slug in slug_used:
            n = 2
            while f"{slug}-{n}" in slug_used: n += 1
            slug = f"{slug}-{n}"
        slug_used.add(slug)
        variants = []
        for v in sorted(vs, key=lambda x: (color_rank(EN_DE_COLOR.get(clean(x[9]).lower(), clean(x[9]))), clean(x[9]))):
            col = EN_DE_COLOR.get(clean(v[9]).lower(), clean(v[9]))
            if v[16] is None: continue
            variants.append({
                "sku": clean(v[1]), "color": col, "ean": clean(v[11]),
                "price": round(float(v[16]) * 0.78, 2), "image": clean(v[37]),
            })
        if not variants: continue
        products.append({
            "sku": sku, "slug": slug, "name": name, "series": serie,
            "category_slug": cat_slug, "short_description": f"Karlik {serie} – {cat_name}",
            "image": variants[0]["image"], "price_min": min(x["price"] for x in variants),
            "warn": False, "variants": variants,
        })
        ame_added += 1

out = {
    "brand": {"name": "Karlik", "slug": "karlik"},
    "categories": [{"name": n, "slug": s} for s, n in sorted(cats.items())],
    "products": products,
}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False)

nv = sum(len(p["variants"]) for p in products)
print(f"Produkte: {len(products)} | Varianten: {nv} | Kategorien: {len(cats)} | ⚠️ Produkte: {len(warn_products)} | American-Style: {ame_added}")
print("Beispiel:", json.dumps(products[0], ensure_ascii=False)[:300])
